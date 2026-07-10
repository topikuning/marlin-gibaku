"""
S-Curve Auto-Generator dari RAB.

Input: RAB parsed dari HPS + kontrak (start/end date)
Output: 
  - Weekly cumulative % target
  - Per-item weekly volume allocation
  - Per-category timeline

Algoritma:
1. Setiap kategori RAB di-map ke "phase window" berdasarkan urutan konstruksi standar KNMP.
   Contoh: Persiapan phase 0-15%, Revetment 10-55%, Landskaping 70-100%.
2. Dalam phase window, distribusikan bobot kategori pakai cubic S-curve (smoothstep).
3. Sum semua kategori per minggu = kurva-S mingguan.
4. Per-item: alokasi volume mingguan proporsional dengan phase kategorinya.

Ini baseline. Bisa di-refine dengan ML kalau data historis banyak.
"""

import json
import math
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Category name keyword → (phase_start_pct, phase_end_pct)
# Based on standard KNMP construction sequence
CATEGORY_PHASE = [
    # (keyword, start%, end%) — first match wins
    ('PERSIAPAN',        0.00, 0.18),
    ('LEVELLING',        0.05, 0.28),
    ('REVETMENT',        0.10, 0.55),
    ('DINDING PENAHAN',  0.15, 0.60),
    ('TAMBATAN',         0.15, 0.60),
    ('DOCKING',          0.20, 0.75),
    ('PONDASI',          0.15, 0.50),
    ('SHELTER PENDARATAN', 0.22, 0.78),
    ('GUDANG BEKU',      0.25, 0.80),
    ('PABRIK ES',        0.30, 0.82),
    ('COOL BOX',         0.30, 0.82),
    ('BENGKEL',          0.30, 0.85),
    ('BALAI NELAYAN',    0.35, 0.85),
    ('KIOS PERBEKALAN',  0.35, 0.85),
    ('SENTRA KULINER',   0.35, 0.85),
    ('PEMASARAN IKAN',   0.35, 0.85),
    ('KANTOR PENGELOLA', 0.35, 0.88),
    ('SHELTER',          0.30, 0.85),  # generic shelter
    ('AREA PARKIR',      0.45, 0.85),
    ('JALAN',            0.40, 0.90),
    ('SALURAN',          0.40, 0.90),
    ('PLUMBING',         0.45, 0.92),
    ('IPAL',             0.50, 0.90),
    ('TPS',              0.55, 0.92),
    ('GENSET',           0.60, 0.92),
    ('PENERANGAN',       0.55, 0.95),
    ('PAGAR',            0.35, 0.90),
    ('GERBANG',          0.55, 0.95),
    ('GAPURA',           0.75, 1.00),
    ('POS JAGA',         0.70, 0.95),
    ('TOILET',           0.60, 0.95),
    ('MUSHOLLA',         0.65, 0.95),
    ('GAZEBO',           0.70, 0.95),
    ('LANDSKAPING',      0.72, 1.00),
]


def get_category_phase(category_name):
    """Match category name to phase window."""
    upper = category_name.upper()
    for kw, start, end in CATEGORY_PHASE:
        if kw in upper:
            return (start, end)
    return (0.25, 0.80)  # default: middle phase


def smoothstep(t):
    """Cubic S-curve: 3t^2 - 2t^3. Smooth start & finish, fast middle."""
    if t <= 0: return 0.0
    if t >= 1: return 1.0
    return 3 * t * t - 2 * t * t * t


def generate_scurve(rab_data, contract_days=150):
    """
    Generate S-curve schedule from RAB.
    
    Args:
        rab_data: dict from parse_hps output
        contract_days: total contract duration in days
    
    Returns:
        dict with weekly targets, category timelines, item schedule
    """
    total_weeks = math.ceil(contract_days / 7)
    grand_total = sum(c['total_value'] for c in rab_data['categories'] if c['total_value'] > 0)

    # Per-week cumulative % planned
    weekly_delta = [0.0] * total_weeks
    category_timelines = []

    # Per-item schedule: (item_ref, week_start, week_end, weekly_volumes)
    item_schedule = []

    for cat in rab_data['categories']:
        if cat['total_value'] <= 0:
            continue
        phase_start_pct, phase_end_pct = get_category_phase(cat['name'])
        week_start = int(phase_start_pct * total_weeks)
        week_end = max(week_start + 1, int(phase_end_pct * total_weeks))
        duration = week_end - week_start

        cat_weight_pct = cat['total_value'] / grand_total * 100

        # Compute weekly progress within category (using smoothstep for S-shape)
        cat_weekly_pcts = []
        prev = 0.0
        for w in range(duration):
            t_now = (w + 1) / duration
            now = smoothstep(t_now)
            cat_weekly_pcts.append(now - prev)
            prev = now

        # Category timeline for viz
        cat_timeline = {
            'roman': cat['roman'],
            'name': cat['name'],
            'value': cat['total_value'],
            'weight_pct': cat_weight_pct,
            'week_start': week_start,
            'week_end': week_end,
            'weekly_delta_pct': [d * cat_weight_pct for d in cat_weekly_pcts],
        }
        category_timelines.append(cat_timeline)

        # Add to global weekly delta
        for i, d in enumerate(cat_weekly_pcts):
            wk = week_start + i
            if 0 <= wk < total_weeks:
                weekly_delta[wk] += d * cat_weight_pct

        # Per-item allocation within category
        def process_items(items, path=''):
            for item in items:
                if item.get('volume') and item.get('unit_price'):
                    vol = item['volume']
                    # Allocate volume across weeks proportional to weekly progress
                    weekly_vols = [vol * d for d in cat_weekly_pcts]
                    item_schedule.append({
                        'ref': f"{cat['roman']}.{item['code']}",
                        'name': item['name'],
                        'category': cat['name'],
                        'total_volume': vol,
                        'unit': item['unit'],
                        'unit_price': item['unit_price'],
                        'total_price': item['total_price'],
                        'week_start': week_start,
                        'week_end': week_end,
                        'weekly_volumes': weekly_vols,
                    })
                if item.get('children'):
                    process_items(item['children'], path + item['code'] + '.')

        process_items(cat['direct_items'])
        for sub in cat['subcategories']:
            process_items(sub['items'])

    # Cumulative
    cumulative = []
    running = 0.0
    for w in weekly_delta:
        running += w
        cumulative.append(min(100.0, running))

    return {
        'total_weeks': total_weeks,
        'contract_days': contract_days,
        'grand_total': grand_total,
        'weekly_delta_pct': weekly_delta,
        'cumulative_pct': cumulative,
        'category_timelines': category_timelines,
        'item_schedule': item_schedule,
    }


def print_scurve_summary(scurve):
    print(f"Total weeks:   {scurve['total_weeks']}")
    print(f"Contract days: {scurve['contract_days']}")
    print(f"Grand total:   Rp {scurve['grand_total']:,.0f}")
    print(f"\nWeekly cumulative plan (Kurva-S):")
    for w in range(0, scurve['total_weeks'], 2):
        c = scurve['cumulative_pct'][w]
        bar = '█' * int(c / 3)
        print(f"  Week {w+1:2d}: {c:5.1f}% {bar}")
    print(f"\nCategory timeline:")
    for c in scurve['category_timelines']:
        print(f"  {c['roman']:5s} W{c['week_start']+1:02d}-W{c['week_end']:02d}  ({c['weight_pct']:5.2f}%)  {c['name'][:50]}")


def export_weekly_plan_excel(scurve, rab_data, out_path):
    """Generate KKP-compatible Weekly Plan / Kurva-S Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kurva-S"

    def border():
        s = Side(style='thin', color='8B7A5C')
        return Border(left=s, right=s, top=s, bottom=s)

    NAVY = '0F2942'
    CREAM = 'F5F0E4'
    CREAM_DARK = 'E8D9BB'
    BRICK = 'C1442E'
    SAGE = '3D7A4F'

    # Header
    ws.merge_cells(f'A1:{chr(65 + scurve["total_weeks"] + 2)}1')
    ws['A1'] = "KURVA-S RENCANA PELAKSANAAN PEKERJAAN"
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{chr(65 + scurve["total_weeks"] + 2)}2')
    ws['A2'] = f"Kampung Nelayan Merah Putih · {rab_data['location']}, {rab_data['province']}"
    ws['A2'].font = Font(name='Arial', size=11)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A3:{chr(65 + scurve["total_weeks"] + 2)}3')
    ws['A3'] = f"Nilai Kontrak: Rp {scurve['grand_total']:,.0f} · Durasi: {scurve['contract_days']} hari ({scurve['total_weeks']} minggu)"
    ws['A3'].font = Font(name='Arial', size=10, italic=True)
    ws['A3'].alignment = Alignment(horizontal='center')

    # Table header
    row = 5
    headers = ['NO', 'URAIAN PEKERJAAN', 'BOBOT (%)']
    for w in range(scurve['total_weeks']):
        headers.append(f'M-{w+1}')
    ws.cell(row=row, column=1).font = Font(bold=True)

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name='Arial', size=9, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = PatternFill('solid', start_color=CREAM_DARK)
        c.border = border()
    ws.row_dimensions[row].height = 25
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    for w in range(scurve['total_weeks']):
        ws.column_dimensions[chr(68 + w)].width = 7 if w < 26 else 5
    row += 1

    # Category rows
    for cat in scurve['category_timelines']:
        ws.cell(row=row, column=1, value=cat['roman']).font = Font(name='Arial', size=9, bold=True)
        ws.cell(row=row, column=2, value=cat['name']).font = Font(name='Arial', size=9, bold=True)
        ws.cell(row=row, column=3, value=round(cat['weight_pct'], 2)).font = Font(name='Arial', size=9, bold=True)
        ws.cell(row=row, column=3).number_format = '0.00'

        for w in range(scurve['total_weeks']):
            val = 0
            idx = w - cat['week_start']
            if 0 <= idx < len(cat['weekly_delta_pct']):
                val = cat['weekly_delta_pct'][idx]
            if val > 0:
                c = ws.cell(row=row, column=4 + w, value=round(val, 3))
                c.number_format = '0.000'
                c.font = Font(name='Arial', size=8)
                # Color intensity based on value
                intensity = min(1.0, val / 3.0)
                red = int(255 - intensity * 60)
                green = int(220 - intensity * 40)
                blue = int(200 - intensity * 40)
                c.fill = PatternFill('solid', start_color=f'{red:02X}{green:02X}{blue:02X}')
            for col in range(1, scurve['total_weeks'] + 4):
                ws.cell(row=row, column=col).border = border()
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        row += 1

    # Empty row
    row += 1

    # Cumulative planned row
    ws.cell(row=row, column=1, value='').font = Font(bold=True)
    ws.cell(row=row, column=2, value='RENCANA KUMULATIF (%)').font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=row, column=3, value=100.00).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=row, column=3).number_format = '0.00'
    for w in range(scurve['total_weeks']):
        c = ws.cell(row=row, column=4 + w, value=round(scurve['cumulative_pct'][w], 2))
        c.font = Font(name='Arial', size=9, bold=True, color=NAVY)
        c.number_format = '0.00'
        c.fill = PatternFill('solid', start_color=CREAM)
        c.alignment = Alignment(horizontal='center')
        c.border = border()
    for col in [1, 2, 3]:
        ws.cell(row=row, column=col).fill = PatternFill('solid', start_color=CREAM)
        ws.cell(row=row, column=col).border = border()
    row += 1

    # Weekly delta row
    ws.cell(row=row, column=1, value='').font = Font()
    ws.cell(row=row, column=2, value='PROGRESS MINGGUAN (%)').font = Font(name='Arial', size=10, bold=True)
    for w in range(scurve['total_weeks']):
        c = ws.cell(row=row, column=4 + w, value=round(scurve['weekly_delta_pct'][w], 2))
        c.font = Font(name='Arial', size=9, color=BRICK)
        c.number_format = '0.00'
        c.alignment = Alignment(horizontal='center')
        c.border = border()
    for col in [1, 2, 3]:
        ws.cell(row=row, column=col).border = border()
    row += 3

    # Signature block
    ws.cell(row=row, column=2, value='Dibuat oleh,').font = Font(size=10)
    ws.cell(row=row, column=8, value='Disetujui oleh,').font = Font(size=10)
    row += 4
    ws.cell(row=row, column=2, value='_______________________').font = Font(size=10)
    ws.cell(row=row, column=8, value='_______________________').font = Font(size=10)
    row += 1
    ws.cell(row=row, column=2, value='Konsultan Perencana').font = Font(size=10, bold=True)
    ws.cell(row=row, column=8, value='Direktorat / PPK').font = Font(size=10, bold=True)
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=scurve['total_weeks'] + 3)
    ws.cell(row=row, column=1, value='Kurva-S auto-generated oleh KNMP Monitor · Algoritma v1.0 · Editable per item oleh PM').font = Font(size=8, italic=True, color='6B6B6B')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

    wb.save(out_path)
    print(f"✓ Kurva-S Excel: {out_path}")


def export_scurve_json(scurve, out_path):
    """Export S-curve data for embedding in prototype."""
    payload = {
        'total_weeks': scurve['total_weeks'],
        'contract_days': scurve['contract_days'],
        'grand_total': scurve['grand_total'],
        'weekly_delta_pct': [round(x, 3) for x in scurve['weekly_delta_pct']],
        'cumulative_pct': [round(x, 2) for x in scurve['cumulative_pct']],
        'category_timelines': [
            {
                'roman': c['roman'],
                'name': c['name'],
                'value': c['value'],
                'weight_pct': round(c['weight_pct'], 2),
                'week_start': c['week_start'],
                'week_end': c['week_end'],
                'weekly_delta_pct': [round(x, 3) for x in c['weekly_delta_pct']],
            } for c in scurve['category_timelines']
        ],
    }
    with open(out_path, 'w') as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    print(f"✓ S-curve JSON: {out_path}")


if __name__ == "__main__":
    # Load parsed Kedungmutih RAB
    with open('/home/claude/knmp/HPS_Kedungmutih.json') as f:
        rab = json.load(f)

    # Assume 150-day contract (~21 weeks)
    scurve = generate_scurve(rab, contract_days=150)

    print_scurve_summary(scurve)
    print(f"\nTotal items scheduled: {len(scurve['item_schedule'])}")

    # Export
    export_weekly_plan_excel(scurve, rab, '/mnt/user-data/outputs/KurvaS-Kedungmutih.xlsx')
    export_scurve_json(scurve, '/home/claude/knmp/scurve_kedungmutih.json')
