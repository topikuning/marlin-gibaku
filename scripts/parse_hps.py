"""
KNMP HPS Parser
Extracts structured RAB (Rencana Anggaran Biaya) from KKP HPS Excel files.

Handles:
- Multi-level hierarchy (Roman numeral categories → sub-categories → items → sub-items)
- Volume + unit for each line item (basis progress tracking)
- Unreliable roman numeral IDs — normalizes using name-based canonical
- Different active scope per location
"""

import re
from openpyxl import load_workbook
from dataclasses import dataclass, field, asdict
from typing import Optional
import json
import sys


@dataclass
class RabItem:
    """A single line item in the RAB with volume + unit."""
    code: str  # e.g. "1", "6.1", "6.1.a"
    name: str
    volume: Optional[float] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    total_price: Optional[float] = None
    tkdn_ratio: Optional[float] = None
    parent_code: Optional[str] = None  # e.g. "6" for "6.1.a"
    children: list = field(default_factory=list)


@dataclass
class RabCategory:
    """Top-level pekerjaan (Roman numeral)."""
    roman: str
    name: str
    total_value: float = 0
    subcategories: list = field(default_factory=list)  # RabSubcategory
    direct_items: list = field(default_factory=list)  # RabItem (if no subcategory)


@dataclass
class RabSubcategory:
    """Sub-level (e.g. III.1, V.2, X.3.1)."""
    code: str  # "III.1"
    name: str
    total_value: float = 0
    items: list = field(default_factory=list)  # RabItem


@dataclass
class LocationRab:
    project: str
    location: str
    province: str
    year: int
    categories: list  # RabCategory
    total: float = 0

    def to_dict(self):
        return {
            "project": self.project,
            "location": self.location,
            "province": self.province,
            "year": self.year,
            "total": self.total,
            "categories": [
                {
                    "roman": c.roman,
                    "name": c.name,
                    "total_value": c.total_value,
                    "subcategories": [
                        {
                            "code": s.code,
                            "name": s.name,
                            "total_value": s.total_value,
                            "items": [_item_dict(it) for it in s.items],
                        } for s in c.subcategories
                    ],
                    "direct_items": [_item_dict(it) for it in c.direct_items],
                } for c in self.categories
            ],
        }


def _item_dict(it):
    d = asdict(it)
    if d.get("children"):
        d["children"] = [_item_dict(c) if isinstance(c, RabItem) else c for c in d["children"]]
    return d


ROMAN = r"^(I{1,3}|IV|V|VI{0,3}|IX|X{1,3}|XIV|XV|XVI{0,3}|XIX|XX{1,3}|XXIV|XXV|XXVI{0,3})$"
SUB_CODE = r"^(I{1,3}|IV|V|VI{0,3}|IX|X{1,3}|XIV|XV|XVI{0,3}|XIX|XX{1,3}|XXIV|XXV|XXVI{0,3})\.\d+\.?$"
ITEM_NUM = r"^\d+$"
SUB_ITEM_NUM = r"^\d+\.\d+\.?$"
SUB_LETTER = r"^[a-z]$"


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_hps(path):
    wb = load_workbook(path, data_only=True)

    # --- Metadata from Resume sheet ---
    project, location, province, year = "", "", "", 0
    if "Resume" in wb.sheetnames:
        for row in wb["Resume"].iter_rows(values_only=True):
            row_str = [str(c).strip() if c is not None else "" for c in row]
            joined = " ".join(row_str)
            if "PROYEK" in joined and ":" in joined:
                # find value after colon
                for i, v in enumerate(row_str):
                    if v == ":" and i + 1 < len(row_str):
                        project = row_str[i + 1]
                        break
            elif "LOKASI" in joined and ":" in joined:
                for i, v in enumerate(row_str):
                    if v == ":" and i + 1 < len(row_str):
                        location = row_str[i + 1]
                        break
            elif "JAWA" in joined or "JATIM" in joined or "JATENG" in joined:
                # province line usually just says "JAWA TENGAH" alone
                for v in row_str:
                    if v.startswith("JAWA") or v in ("BANTEN", "DIY", "LAMPUNG", "BENGKULU"):
                        province = v
                        break
            elif "TAHUN ANGGARAN" in joined and ":" in joined:
                for i, v in enumerate(row_str):
                    if v == ":" and i + 1 < len(row_str):
                        try:
                            year = int(row_str[i + 1])
                        except ValueError:
                            pass
                        break

    # --- Category totals from Sub Resume ---
    category_totals = {}  # roman → (name, value)
    subcategory_totals = {}  # code (e.g. "V.1") → (name, value)

    if "Sub Resume" in wb.sheetnames:
        ws = wb["Sub Resume"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not cells or not cells[0]:
                continue
            code = cells[0].rstrip('.').strip()
            # Roman numeral (category)
            if re.match(ROMAN, code):
                name = ""
                value = None
                for c in cells[1:]:
                    if c and not name:
                        name = c
                        continue
                    n = _num(c)
                    if n is not None and n > 100:  # rupiah scale
                        value = n
                        break
                if name and value is not None:
                    category_totals[code] = (name, value)
            # Sub-code (V.1, X.3.1)
            elif "." in code and re.match(ROMAN + r"\.\d+", code.split('.')[0] + '.' + code.split('.')[1]) if len(code.split('.')) >= 2 else False:
                pass  # handled below more permissively

            # More permissive sub-code detection
            m = re.match(r"^([IVX]+)\.(\d+(?:\.\d+)*)\.?$", code)
            if m:
                name = ""
                value = None
                for c in cells[1:]:
                    if c and not name:
                        name = c
                        continue
                    n = _num(c)
                    if n is not None and n > 100:
                        value = n
                        break
                if name:
                    subcategory_totals[code.rstrip('.')] = (name, value or 0)

    # --- Detail items from RAB sheet ---
    categories = []
    current_cat = None
    current_sub = None
    current_item = None

    if "RAB" in wb.sheetnames:
        ws = wb["RAB"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not cells or not any(cells):
                continue
            first = cells[0]
            # Second column = name/description
            name_col = cells[1] if len(cells) > 1 else ""

            # Roman numeral = new category
            if re.match(ROMAN, first) and name_col and name_col.startswith("PEKERJAAN"):
                if current_cat:
                    categories.append(current_cat)
                total = category_totals.get(first, (name_col, 0))[1]
                current_cat = RabCategory(roman=first, name=name_col, total_value=total)
                current_sub = None
                continue

            # Sub-category code (III.1, V.2, X.3.1)
            m_sub = re.match(r"^([IVX]+)\.(\d+(?:\.\d+)*)\.?$", first)
            if m_sub and current_cat and name_col and name_col.startswith("Pekerjaan"):
                sub_code = first.rstrip('.')
                total = subcategory_totals.get(sub_code, (name_col, 0))[1]
                current_sub = RabSubcategory(code=sub_code, name=name_col, total_value=total)
                current_cat.subcategories.append(current_sub)
                continue

            # Item number (1, 2, 3) — main item
            if re.match(ITEM_NUM, first) and current_cat and name_col:
                vol = _num(cells[4]) if len(cells) > 4 else None
                unit = cells[5] if len(cells) > 5 and vol is not None else None
                unit_price = _num(cells[6]) if len(cells) > 6 else None
                total_price = _num(cells[7]) if len(cells) > 7 else None
                tkdn = _num(cells[8]) if len(cells) > 8 else None

                item = RabItem(
                    code=first,
                    name=name_col,
                    volume=vol,
                    unit=unit,
                    unit_price=unit_price,
                    total_price=total_price,
                    tkdn_ratio=tkdn,
                )
                current_item = item
                if current_sub:
                    current_sub.items.append(item)
                else:
                    current_cat.direct_items.append(item)
                continue

            # Sub-item (6.1, 6.2, 6.1.a) — with letter suffix or dotted number
            m_subnum = re.match(r"^(\d+)\.(\d+)\.?$", first)
            m_letter = re.match(SUB_LETTER, first)

            if (m_subnum or m_letter) and current_item and name_col:
                vol = _num(cells[4]) if len(cells) > 4 else None
                unit = cells[5] if len(cells) > 5 and vol is not None else None
                unit_price = _num(cells[6]) if len(cells) > 6 else None
                total_price = _num(cells[7]) if len(cells) > 7 else None

                sub_item = RabItem(
                    code=f"{current_item.code}.{first}",
                    name=name_col,
                    volume=vol,
                    unit=unit,
                    unit_price=unit_price,
                    total_price=total_price,
                    parent_code=current_item.code,
                )
                current_item.children.append(sub_item)

    if current_cat:
        categories.append(current_cat)

    # Compute total
    total = sum(c.total_value for c in categories)

    return LocationRab(
        project=project,
        location=location,
        province=province,
        year=year,
        categories=categories,
        total=total,
    )


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/project/HPS_Kedungmutih.xlsx"
    rab = parse_hps(path)
    print(f"Project: {rab.project}")
    print(f"Location: {rab.location}, {rab.province} ({rab.year})")
    print(f"Total: Rp {rab.total:,.0f}")
    print(f"Categories: {len(rab.categories)}")
    print()

    active_cats = [c for c in rab.categories if c.total_value > 0]
    print(f"Active categories: {len(active_cats)}")
    for c in active_cats[:5]:
        print(f"  {c.roman}. {c.name} — Rp {c.total_value:,.0f}")
        item_count = sum(len(s.items) for s in c.subcategories) + len(c.direct_items)
        sub_item_count = sum(
            sum(len(i.children) for i in s.items) for s in c.subcategories
        ) + sum(len(i.children) for i in c.direct_items)
        print(f"     └─ {len(c.subcategories)} sub · {item_count} items · {sub_item_count} sub-items")

    total_items = sum(
        sum(len(s.items) for s in c.subcategories) + len(c.direct_items)
        for c in rab.categories
    )
    total_sub_items = sum(
        sum(sum(len(i.children) for i in s.items) for s in c.subcategories)
        + sum(len(i.children) for i in c.direct_items)
        for c in rab.categories
    )
    print(f"\nTotal line items: {total_items} (+ {total_sub_items} sub-items)")

    # Save as JSON
    out_path = path.replace('.xlsx', '.json').replace('/mnt/project/', '/home/claude/knmp/')
    with open(out_path, 'w') as f:
        json.dump(rab.to_dict(), f, indent=2, ensure_ascii=False)
    print(f"\nWritten: {out_path}")
